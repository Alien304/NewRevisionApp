from distutils.dir_util import copy_tree
import os
import glob
import openpyxl
from datetime import date
import customtkinter
from tkinter import *

def Creating_doc():
    Project_car = Project_car_combo.get()
    Project_A_or_B = Project_A_or_B_combo.get()
    Project_batch_old = Project_batch_old_en.get()
    Project_batch_new = Project_batch_new_en.get()
    change_from_wl = change_from_wl_en.get()
    change_to_wl = change_to_wl_en.get()
    Project_harnesses = Project_harnesses_combo.get()
            
    if Project_car == "B_AB":
        change_from_wl_car = str(f"{Project_A_or_B}_{change_from_wl}")
        change_to_wl_car = str(f"{Project_A_or_B}_{change_to_wl}")
        copy_from_destination = str("Outcome\CAR " + Project_car + "\\" + Project_batch_old + "\Harnesses " + Project_harnesses + "\\" + Project_A_or_B + "\\" + change_from_wl)
        copy_to_destination = str("Outcome\CAR " + Project_car + "\\" + Project_batch_new + "\Harnesses " + Project_harnesses + "\\" + Project_A_or_B + "\\" + change_to_wl)
        copy_tree(copy_from_destination, copy_to_destination)
    else:
        change_from_wl_car = str(f"{Project_car}_{change_from_wl}")
        change_to_wl_car =str(f"{Project_car}_{change_to_wl}")
        copy_from_destination = str("Outcome\CAR " + Project_car + "\\" + Project_batch_old + "\Harnesses " + Project_harnesses + "\\" + change_from_wl)
        copy_to_destination = str("Outcome\CAR " + Project_car + "\\" + Project_batch_new + "\Harnesses " + Project_harnesses + "\\" + change_to_wl)
        copy_tree(copy_from_destination, copy_to_destination)
    print(f"Zakończono tworzenie folderu {copy_to_destination}")
    
    for file in glob.iglob(f"{copy_to_destination}\**\*.xlsx", recursive=True):
        if file.endswith(f"{change_from_wl_car}.xlsx"):
            new_file = file.replace(change_from_wl_car, change_to_wl_car)
            os.rename(file, new_file)
                    
    for folder in glob.iglob(f"{copy_to_destination}\**\*", recursive=True):
        if folder.endswith(f"{change_from_wl_car}"):
            new_folder = folder.replace(change_from_wl_car, change_to_wl_car)
            os.rename(folder, new_folder)

        
def Generate_change():
    CAR_NAME = CAR_NAME_combo.get()
    sheet_names = ['Cutting List', 'Connection List', '1 z 13', '2 z 13', '3 z 13',
                '4 z 13', '5 z 13', '6 z 13', '7 z 13', '8 z 13', '9 z 13', 
                '10 z 13', '11 z 13', '12 z 13', '13 z 13', '1 z 12', '2 z 12', '3 z 12',
                '4 z 12', '5 z 12', '6 z 12', '7 z 12', '8 z 12', '9 z 12', 
                '10 z 12', '11 z 12', '12 z 12']
    LC_path = []
    LP_path = []
    Old = []
    New = []
    def generate_info():
        changefromto = openpyxl.load_workbook("Change_from_to.xlsx")
        for name in changefromto.sheetnames:
            sheet1 = changefromto['Arkusz1']
            for column in sheet1.iter_cols(min_col=1, max_col=4, min_row=2):
                for a in column:
                    Old.append(a.value)       
                for column2 in sheet1.iter_cols(min_col=5, max_col=8, min_row=2):
                    for b in column2:
                        New.append(b.value)
                for column3 in sheet1.iter_cols(min_col=9, min_row=2, max_col=9):
                    for c in column3:
                        LC_path.append(c.value)
                for column4 in sheet1.iter_cols(min_col=10, min_row=2, max_col=10):
                    for d in column4:
                        LP_path.append(d.value)

    generate_info()
                
    replacement_pair = dict(zip(Old, New))

    def cutting_list():
        try:
            excel = openpyxl.load_workbook(save_path_lc)
            for name in excel.sheetnames:
                if name in sheet_names:      
                    sheet = excel[name]
                    for row in sheet.iter_rows(min_col=6, max_col=6):
                        for cell in row:
                            if cell.value in replacement_pair.keys():
                                cell.value = replacement_pair.get(cell.value)
                    for dates_row in sheet.iter_rows(min_col=11, max_col=11, min_row=4, max_row=6):
                        for dates in dates_row:
                            dates.value = date.today()
            excel.save(save_path_lc)
        except:
            print(f"File {save_path_lc} does not exist.")
            pass
        
    def connection_list():
        try:    
            excel = openpyxl.load_workbook(save_path_lp)
            for name in excel.sheetnames:
                if name in sheet_names:      
                    sheet = excel[name]
                    for row in sheet.iter_rows(min_col=9, max_col=17):
                        for cell in row:
                            if cell.value in replacement_pair.keys():
                                cell.value = replacement_pair.get(cell.value)
                    for dates_row in sheet.iter_rows(min_col=15, max_col=15, min_row=22, max_row=24):
                        for dates in dates_row:
                            dates.value = date.today()
            excel.save(save_path_lp)
        except:
            print(f"File {save_path_lp} does not exist.")
            pass    

    if CAR_NAME == "Bx":
        for i_lc in LC_path[0:26]:
            save_path_lc = i_lc
            cutting_list()
        for i_lp in LP_path[0:26]:
            save_path_lp = i_lp
            connection_list()
            
    elif CAR_NAME == "BDx":
        for i_lc in LC_path[27:53]:
            save_path_lc = i_lc
            cutting_list()
        for i_lp in LP_path[27:53]:
            save_path_lp = i_lp
            connection_list()

    elif CAR_NAME == "B":
        for i_lc in LC_path[54:82]:
            save_path_lc = i_lc
            cutting_list()
        for i_lp in LP_path[54:82]:
            save_path_lp = i_lp
            connection_list()

    elif CAR_NAME == "AB":
        for i_lc in LC_path[83:111]:
            save_path_lc = i_lc
            cutting_list()
        for i_lp in LP_path[83:111]:
            save_path_lp = i_lp
            connection_list()

    elif CAR_NAME == "BD":
        for i_lc in LC_path[112:136]:
            save_path_lc = i_lc
            cutting_list()
        for i_lp in LP_path[112:136]:
            save_path_lp = i_lp
            connection_list()

    else:
        print("Script end because of wrong value")  

    print("Zakończono")

customtkinter.set_appearance_mode("dark")
customtkinter.set_default_color_theme("dark-blue")
root = customtkinter.CTk()
root.geometry("850x850")
root.title("Project - Aktualizacja dokumentacji")

frame_a = customtkinter.CTkFrame(master=root)
frame_a.pack(pady=20, padx=60, side=LEFT)

label = customtkinter.CTkLabel(master=frame_a, justify="center",
                                text="Wybierz rodzaj wagonu:")
label.pack(pady=2, padx=50)
Project_car_combo = customtkinter.CTkComboBox(master=frame_a, justify="center",
                                   values = ["-", "B_AB", "BDx", "BD", "Bx"])
Project_car_combo.pack(pady=20, padx=20)

label = customtkinter.CTkLabel(master=frame_a, justify="center",
                               text="Sprecyzuj wagon B_AB, A lub B:")
label.pack(pady=2, padx=50)
Project_A_or_B_combo = customtkinter.CTkComboBox(master=frame_a, justify="center",
                                   values = ["-", "B", "AB"])
Project_A_or_B_combo.pack(pady=20, padx=20)

label = customtkinter.CTkLabel(master=frame_a, width=200, justify="center",
                                text="Wpisz akutalną wersję Batch (Np. Batch 9):")
label.pack(pady=2, padx=50)
Project_batch_old_en = customtkinter.CTkEntry(master=frame_a, width=200, justify="center")
Project_batch_old_en.pack(pady=20, padx=20)

label = customtkinter.CTkLabel(master=frame_a, width=200, justify="center",
                               text="Wpisz docelową wersję Batch (Np. Batch 10):")
label.pack(pady=2, padx=50)
Project_batch_new_en = customtkinter.CTkEntry(master=frame_a, width=200, justify="center")
Project_batch_new_en.pack(pady=20, padx=20)

label = customtkinter.CTkLabel(master=frame_a, width=200, justify="center",
                               text="Wpisz aktualną wersję WL (np. WL_9-0_M00):")
label.pack(pady=2, padx=50)
change_from_wl_en = customtkinter.CTkEntry(master=frame_a, width=200, justify="center")
change_from_wl_en.pack(pady=20, padx=20)

label = customtkinter.CTkLabel(master=frame_a, width=200, justify="center", 
                               text="Wpisz docelową wersję WL (np. WL_9-0_M01):")
label.pack(pady=2, padx=50)
change_to_wl_en = customtkinter.CTkEntry(master=frame_a, width=200, justify="center")
change_to_wl_en.pack(pady=20, padx=20)

label = customtkinter.CTkLabel(master=frame_a, width=200, justify="center",
                               text="Wybierz rodzaj wiązek do kopiowania:")
label.pack(pady=2, padx=50)
Project_harnesses_combo = customtkinter.CTkComboBox(master=frame_a, width=200, justify="center",
                                   values = ["-", "BAT", "CAB", "CAB EX", "CAB RF", "DOORS", "EX", "IN", "RF"])
Project_harnesses_combo.pack(pady=20, padx=50)

create_button = customtkinter.CTkButton(master=frame_a, text="Create", command=Creating_doc)
create_button.pack(pady=20, padx=20)

frame_b = customtkinter.CTkFrame(master=root)
frame_b.pack(pady=20, padx=60, side=RIGHT)

label = customtkinter.CTkLabel(master=frame_b, width=200, justify="center",
                                text="Wybierz rodzaj wagonu:")
label.pack(pady=5, padx=50)
CAR_NAME_combo = customtkinter.CTkComboBox(master=frame_b, justify="center", 
                                   values = ["-","AB","B","BDx","BD","Bx"])
CAR_NAME_combo.pack(pady=20, padx=20)

change_button = customtkinter.CTkButton(master=frame_b, text="Change", command=Generate_change)
change_button.pack(pady=20, padx=20)
    
root.mainloop()